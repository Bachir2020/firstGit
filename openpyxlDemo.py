import openpyxl

book = openpyxl.load_workbook("C:\\Bachir\\Formations\\Selenium_Python\\e2eFramework\\testData\\openxlDemoFile.xlsx")#permet de charger le ficher excel
sheet = book.active #se pointer sur l'onglet actif

Dic={}
for i in range (1,sheet.max_row+1):
    if sheet.cell(row=i,column=1).value=="test33":
        for j in range(2,sheet.max_column+1):

            Dic[sheet.cell(row=1,column=j).value] = sheet.cell(row=i,column=j).value

print(Dic)