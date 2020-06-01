import openpyxl


class HomePageData:
    homePageData= {"name":"bachir","mail":"@atm", "password":"2011" , "gender":"Male"},{"name":"wafa","mail":"@salhi", "password":"2020" , "gender":"Female"}

    @staticmethod #ainsi on a pas besoin de créer un objet pour appeler cette méthode dans les autres fichiers
    #et donc pas besoin de passer self en paramètre :)

    def getHomePageData(testName):

        book = openpyxl.load_workbook(
            "C:\\Bachir\\Formations\\Selenium_Python\\e2eFramework\\testData\\openxlDemoFile.xlsx")  # permet de charger le ficher excel
        sheet = book.active  # se pointer sur l'onglet actif

        Dic = {}
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == testName:
                for j in range(2, sheet.max_column + 1):
                    Dic[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return [Dic] # pour envoyer le dictionnaire sous forme de liste. Car la fixture dans le fichier test_homePage.py ne prend en paramètre que le type liste